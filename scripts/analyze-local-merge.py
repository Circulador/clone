"""Analisa merge local das 3 planilhas baixadas pelo usuário."""
import openpyxl
import re
from datetime import datetime

FILES = {
    'B': r'C:\Users\81008280\Downloads\B APOS ENTREGA NO BALCAO ATE EMAIL  JA SOU ITALIANO (10).xlsx',
    'JSI': r'C:\Users\81008280\Downloads\JSI apos receber e mail Ja sou Italiano aguardando finalizaco consular RJ (1).xlsx',
    'P': r'C:\Users\81008280\Downloads\P Espera por Pendencias Tempo Consulado SEDEX (1).xlsx',
}
STAGE = {'P': 3, 'JSI': 2, 'B': 1}
BAD = re.compile(r'^(MEDIA|MÉDIA|MENOR|MAIOR|TOTAL|ULTIMO|ÚLTIMO|ESTAMOS|DATA DE|NOME|RECEBIDO|GRUPO|N[ºo°]|B PLAN|PLANILHA|ATUALIZAD|DIAS)', re.I)

def iso_cell(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None

def layout(pid):
    if pid == 'P':
        return 1, 0, 4, 2
    return 0, 1, 4, 2

def parse_file(pid, path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    ec, nc, rc, gc = layout(pid)
    recs = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= max(ec, nc):
            continue
        entrega = iso_cell(row[ec])
        nome = str(row[nc]).strip() if row[nc] else ''
        if not entrega or not nome or BAD.match(nome):
            continue
        resol = iso_cell(row[rc]) if rc < len(row) else None
        grupo = str(row[gc]).strip() if gc < len(row) and row[gc] else '-'
        recs.append({'entrega': entrega, 'nome': nome, 'grupo': grupo, 'resol': resol, 'planilha': pid})
    wb.close()
    return recs

def norm(n):
    return ' '.join(n.upper().split())

def pick(a, b):
    pa, pb = STAGE[a['planilha']], STAGE[b['planilha']]
    if pa != pb:
        return a if pa > pb else b
    if bool(a['resol']) != bool(b['resol']):
        return b if a['resol'] else a
    return a if a['entrega'] >= b['entrega'] else b

def merge(records):
    by = {}
    dropped = []
    for r in records:
        k = norm(r['nome'])
        if k not in by:
            by[k] = r
        else:
            kept = pick(by[k], r)
            rem = r if kept is by[k] else by[k]
            dropped.append((k, kept['planilha'], rem['planilha']))
            by[k] = kept
    return list(by.values()), dropped

all_recs = []
for pid, path in FILES.items():
    try:
        recs = parse_file(pid, path)
        print(f'{pid}: {len(recs)} brutos')
        all_recs.extend(recs)
    except FileNotFoundError:
        print(f'{pid}: arquivo não encontrado')

raw = len(all_recs)
merged, dropped = merge(all_recs)
print(f'\nBruto: {raw}')
print(f'Únicos: {len(merged)}')
print(f'Removidos: {len(dropped)}')
from collections import Counter
print('Transições:', dict(Counter(f'{r}→{k}' for _, k, r in dropped)))
for k, kept, rem in dropped[:20]:
    print(f'  {k}: {rem} removido, {kept} mantido')
